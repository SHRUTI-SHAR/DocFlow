"""
Excel Service - Parse and generate Excel files
"""

import logging
from typing import List, Dict, Any, Optional, BinaryIO
from io import BytesIO
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import csv

logger = logging.getLogger(__name__)


class ExcelService:
    """Service for parsing and generating Excel files"""
    
    @staticmethod
    def get_worksheets(file_content: bytes, filename: str) -> List[Dict[str, Any]]:
        """
        Get list of all worksheets in an Excel file with their column counts
        
        Args:
            file_content: Raw bytes of the Excel file
            filename: Original filename
            
        Returns:
            List of worksheets with name, column_count, row_count
        """
        try:
            # Don't use read_only mode - it can miss cells
            workbook = openpyxl.load_workbook(BytesIO(file_content), read_only=False, data_only=True)
            
            worksheets = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Get the actual dimensions of the sheet
                max_col = sheet.max_column or 1
                
                # Read ALL columns from first row up to max_column
                columns = []
                for col_idx in range(1, max_col + 1):
                    cell_value = sheet.cell(row=1, column=col_idx).value
                    if cell_value is not None and str(cell_value).strip():
                        columns.append(str(cell_value).strip())
                
                # Count data rows
                max_row = sheet.max_row or 1
                row_count = max(0, max_row - 1)  # Exclude header row
                
                worksheets.append({
                    "name": sheet_name,
                    "column_count": len(columns),
                    "row_count": row_count,
                    "columns": columns[:10]  # Preview first 10 columns
                })
            
            workbook.close()
            
            logger.info(f"📊 Found {len(worksheets)} worksheets in '{filename}'")
            return worksheets
            
        except Exception as e:
            logger.error(f"❌ Failed to get worksheets: {e}")
            raise ValueError(f"Failed to read Excel file: {str(e)}")
    
    @staticmethod
    def parse_template(file_content: bytes, filename: str, sheet_name: str = None) -> Dict[str, Any]:
        """
        Parse an Excel template file and extract ALL column headers from specified sheet
        
        Args:
            file_content: Raw bytes of the Excel file
            filename: Original filename
            sheet_name: Specific sheet to parse (default: first sheet)
            
        Returns:
            Dict with columns, sheet_name, row_count, all_sheets
        """
        try:
            # Don't use read_only mode - it can miss cells in sparse sheets
            workbook = openpyxl.load_workbook(BytesIO(file_content), read_only=False, data_only=True)
            
            # Get all sheet names
            all_sheets = workbook.sheetnames
            
            # Use specified sheet or first sheet
            if sheet_name and sheet_name in all_sheets:
                target_sheet = sheet_name
            else:
                target_sheet = all_sheets[0]
            
            sheet = workbook[target_sheet]
            
            # Get actual dimensions
            max_col = sheet.max_column or 1
            max_row = sheet.max_row or 1
            
            logger.info(f"📊 Sheet '{target_sheet}' dimensions: {max_col} columns x {max_row} rows")
            
            # Extract ALL headers from first row by iterating through all columns
            columns = []
            for col_idx in range(1, max_col + 1):
                cell_value = sheet.cell(row=1, column=col_idx).value
                if cell_value is not None and str(cell_value).strip():
                    columns.append(str(cell_value).strip())
            
            # Count data rows (excluding header)
            row_count = max(0, max_row - 1)
            
            workbook.close()
            
            logger.info(f"📊 Parsed Excel template '{filename}' sheet '{target_sheet}': {len(columns)} columns, {row_count} rows")
            
            return {
                "columns": columns,
                "sheet_name": target_sheet,
                "row_count": row_count,
                "all_sheets": all_sheets
            }
            
        except Exception as e:
            logger.error(f"❌ Failed to parse Excel template: {e}")
            raise ValueError(f"Failed to parse Excel file: {str(e)}")
    
    @staticmethod
    def generate_excel(
        columns: List[str],
        rows: List[Dict[str, Any]],
        sheet_name: str = "Extracted Data"
    ) -> bytes:
        """
        Generate an Excel file with mapped data
        
        Args:
            columns: List of column headers
            rows: List of row data (each row is dict: {column: value})
            sheet_name: Name for the worksheet
            
        Returns:
            Excel file as bytes
        """
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col_idx, column in enumerate(columns, 1):
            cell = sheet.cell(row=1, column=col_idx, value=column)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write data rows
        logger.info(f"📝 Writing {len(rows)} rows to Excel")
        values_written = 0
        empty_values = 0
        
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, column in enumerate(columns, 1):
                value = row_data.get(column, "")
                if value and value != "":
                    values_written += 1
                else:
                    empty_values += 1
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border
        
        logger.info(f"✅ Excel write complete: {values_written} values written, {empty_values} empty cells")
        
        # Auto-adjust column widths
        for col_idx, column in enumerate(columns, 1):
            max_length = len(column)
            
            for row_idx in range(2, len(rows) + 2):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    # Limit to first 50 chars for width calculation
                    max_length = max(max_length, min(len(str(cell_value)), 50))
            
            # Set width with some padding
            adjusted_width = min(max_length + 2, 60)
            sheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        
        # Freeze header row
        sheet.freeze_panes = "A2"
        
        # Save to bytes
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        logger.info(f"📄 Generated Excel file: {len(columns)} columns, {len(rows)} rows")
        
        return output.getvalue()
    
    @staticmethod
    def generate_csv(
        columns: List[str],
        rows: List[Dict[str, Any]]
    ) -> bytes:
        """
        Generate a CSV file with mapped data
        
        Args:
            columns: List of column headers
            rows: List of row data (each row is dict: {column: value})
            
        Returns:
            CSV file as bytes (UTF-8 encoded)
        """
        output = BytesIO()
        
        # Write with UTF-8 BOM for Excel compatibility
        output.write(b'\xef\xbb\xbf')
        
        # Create CSV writer
        import codecs
        wrapper = codecs.getwriter('utf-8')(output)
        writer = csv.DictWriter(wrapper, fieldnames=columns, extrasaction='ignore')
        
        # Write header
        writer.writeheader()
        
        # Write rows
        for row in rows:
            # Ensure all values are strings
            cleaned_row = {k: str(v) if v is not None else "" for k, v in row.items()}
            writer.writerow(cleaned_row)
        
        logger.info(f"📄 Generated CSV file: {len(columns)} columns, {len(rows)} rows")
        
        return output.getvalue()
    
    @staticmethod
    def validate_file(filename: str) -> bool:
        """Check if file is a valid Excel file"""
        valid_extensions = ['.xlsx', '.xls', '.xlsm']
        return any(filename.lower().endswith(ext) for ext in valid_extensions)
