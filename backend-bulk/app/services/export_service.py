"""
Export Service - Generate CSV/Excel files from granular field data
Queries bulk_extracted_fields table and builds structured exports
"""

import logging
from typing import List, Dict, Any, Optional
from io import BytesIO, StringIO
import csv
import asyncpg

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ..core.config import settings

logger = logging.getLogger(__name__)


class ExportService:
    """
    Handles CSV and Excel export generation from bulk_extracted_fields
    Supports multiple export formats: summary, detailed, pivoted
    """
    
    def __init__(self):
        self.pool: Optional[asyncpg.Pool] = None
        logger.info("📊 ExportService initialized")
    
    async def initialize_pool(self):
        """Create database connection pool"""
        if self.pool is not None:
            return
        
        self.pool = await asyncpg.create_pool(
            settings.DATABASE_URL,
            min_size=1,
            max_size=5,  # Transaction pooler handles actual pooling
            max_inactive_connection_lifetime=60,
            statement_cache_size=0,  # CRITICAL: Disable prepared statements for pgbouncer
        )
        logger.info("✅ Export service pool created (transaction pooler mode)")
    
    async def close_pool(self):
        """Close connection pool"""
        if self.pool:
            await self.pool.close()
            self.pool = None
    
    async def query_job_fields(self, job_id: str) -> List[Dict[str, Any]]:
        """
        Query all fields for a job, organized by document
        
        Returns:
            List of documents with their extracted fields
        """
        if not self.pool:
            await self.initialize_pool()
        
        async with self.pool.acquire() as conn:
            # Query all fields with document info
            rows = await conn.fetch("""
                SELECT 
                    d.id as document_id,
                    d.filename as document_name,
                    d.source_path,
                    f.field_name,
                    f.field_label,
                    f.field_type,
                    f.field_value,
                    f.confidence_score,
                    f.page_number,
                    f.validation_status,
                    f.needs_manual_review
                FROM bulk_job_documents d
                JOIN bulk_extracted_fields f ON d.id = f.document_id
                WHERE d.job_id = $1 
                  AND d.status = 'completed'
                ORDER BY d.filename, f.page_number, f.field_name
            """, job_id)
            
            # Group by document
            documents = {}
            for row in rows:
                doc_id = str(row["document_id"])
                if doc_id not in documents:
                    documents[doc_id] = {
                        "document_id": doc_id,
                        "document_name": row["document_name"],
                        "source_path": row["source_path"],
                        "fields": []
                    }
                
                documents[doc_id]["fields"].append({
                    "name": row["field_name"],
                    "label": row["field_label"],
                    "type": row["field_type"],
                    "value": row["field_value"],
                    "confidence": float(row["confidence_score"]) if row["confidence_score"] else 0.0,
                    "page": row["page_number"],
                    "status": row["validation_status"],
                    "needs_review": row["needs_manual_review"]
                })
            
            logger.info(f"📊 Retrieved {len(rows)} fields from {len(documents)} documents")
            return list(documents.values())
    
    async def generate_csv_export(self, job_id: str) -> bytes:
        """
        Generate CSV export (flat format)
        One row per field across all documents
        """
        documents = await self.query_job_fields(job_id)
        
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=[
            "document_name",
            "field_name",
            "field_label",
            "field_type",
            "field_value",
            "confidence",
            "page",
            "validation_status",
            "needs_review"
        ])
        
        writer.writeheader()
        
        for doc in documents:
            for field in doc["fields"]:
                writer.writerow({
                    "document_name": doc["document_name"],
                    "field_name": field["name"],
                    "field_label": field["label"],
                    "field_type": field["type"],
                    "field_value": field["value"],
                    "confidence": field["confidence"],
                    "page": field["page"],
                    "validation_status": field["status"],
                    "needs_review": field["needs_review"]
                })
        
        csv_bytes = output.getvalue().encode("utf-8")
        logger.info(f"✅ Generated CSV export ({len(csv_bytes)} bytes)")
        return csv_bytes
    
    async def generate_excel_summary(self, job_id: str) -> bytes:
        """
        Generate Excel with summary format
        One row per document with key statistics
        """
        documents = await self.query_job_fields(job_id)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Document Summary"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Headers
        headers = ["Document Name", "Total Fields", "Avg Confidence", "Pages", "Needs Review"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row_num, doc in enumerate(documents, 2):
            fields = doc["fields"]
            avg_confidence = sum(f["confidence"] for f in fields) / len(fields) if fields else 0
            max_page = max((f["page"] for f in fields), default=0)
            review_count = sum(1 for f in fields if f["needs_review"])
            
            ws.cell(row=row_num, column=1, value=doc["document_name"])
            ws.cell(row=row_num, column=2, value=len(fields))
            ws.cell(row=row_num, column=3, value=round(avg_confidence, 3))
            ws.cell(row=row_num, column=4, value=max_page)
            ws.cell(row=row_num, column=5, value=review_count)
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Save to bytes
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        excel_bytes = buffer.getvalue()
        logger.info(f"✅ Generated Excel summary ({len(excel_bytes)} bytes)")
        return excel_bytes
    
    async def generate_excel_pivoted(self, job_id: str) -> bytes:
        """
        Generate Excel in pivoted format (like BNI)
        Columns = field names, Rows = documents
        """
        documents = await self.query_job_fields(job_id)
        
        # Get all unique field names
        all_field_names = set()
        for doc in documents:
            for field in doc["fields"]:
                all_field_names.add(field["name"])
        
        sorted_fields = sorted(all_field_names)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Fields by Document"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Headers: Document Name | field1 | field2 | ... | fieldN
        headers = ["Document Name"] + sorted_fields
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        
        # Data rows
        for row_num, doc in enumerate(documents, 2):
            # Create field lookup
            field_lookup = {f["name"]: f["value"] for f in doc["fields"]}
            
            # Document name in first column
            ws.cell(row=row_num, column=1, value=doc["document_name"])
            
            # Fill in field values
            for col_num, field_name in enumerate(sorted_fields, 2):
                value = field_lookup.get(field_name, "")
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Auto-size columns (with limits)
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 40)
        
        # Save to bytes
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        excel_bytes = buffer.getvalue()
        logger.info(f"✅ Generated Excel pivoted ({len(excel_bytes)} bytes)")
        return excel_bytes


# Singleton instance
_export_service: Optional[ExportService] = None

async def get_export_service() -> ExportService:
    """Get or create export service singleton"""
    global _export_service
    if _export_service is None:
        _export_service = ExportService()
        await _export_service.initialize_pool()
    return _export_service
